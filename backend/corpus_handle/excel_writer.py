import openpyxl,os


def get_score_of_context(context):
    DEFAUT_GROUP_MEMBER_NUM = 10
    length = len(context)
    group_num = length // DEFAUT_GROUP_MEMBER_NUM  + 1
    return group_num*100

def get_threshold_value_of_context(context):
    return 80

def writeOneExcel(lis,filename):
    MIN_LEN = 5
    f = openpyxl.Workbook()  # 创建工作簿
    sheet1 = f.active
    i = 1
    for c in lis:
        if len(c) >= MIN_LEN:
            score = get_score_of_context(c)
            threshold_value = get_threshold_value_of_context(c)
            sheet1.cell(row=i, column=1).value = c
            sheet1.cell(row=i, column=2).value = score
            sheet1.cell(row=i, column=3).value = threshold_value
            i = i+1
    f.save(filename)

##################################
# BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# excels_dir = os.path.join(BASE_DIR, 'corpus_handle/excels/')
# save_dir = excels_dir + '1.xlsx'
# print(save_dir)
#
#
# f = openpyxl.Workbook() #创建工作簿
# sheet1 = f.active
#
# sheet1.cell(row=1,column=1).value = 10
#
# f.save(save_dir)